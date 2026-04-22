# ============================================================
# excel_reporter.py - 분석 결과를 Excel(.xlsx) 파일로 저장
# ============================================================
# openpyxl 라이브러리를 사용해서 분석 결과를 보기 좋은
# Excel 파일로 만듭니다. 시트별로 색상과 서식을 구분합니다.
# ============================================================

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional

# openpyxl: Python에서 Excel 파일을 만들고 편집하는 라이브러리
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from models.log_entry import (
    AbnormalAccessRecord,
    ApplicationFlowRecord,
    ErrorRecord,
    ResponseTimeRecord,
    SpikeRecord,
    WalletAnomalyRecord,
)


# ============================================================
# 색상 코드 상수 (ARGB 형식: Alpha+RGB 6자리 16진수)
# ============================================================
COLOR_HEADER_BLUE   = "FF2E75B6"  # 헤더 배경 (진한 파란색)
COLOR_HEADER_LIGHT  = "FFDCE6F1"  # 보조 헤더 배경 (연한 파란색)
COLOR_RED_FILL      = "FFFFC7CE"  # 심각한 이상 행 배경 (연한 빨간색)
COLOR_ORANGE_FILL   = "FFFFEB9C"  # 경고 행 배경 (연한 주황색)
COLOR_GREEN_FILL    = "FFC6EFCE"  # 정상 행 배경 (연한 초록색)
COLOR_WHITE         = "FFFFFFFF"  # 흰색
COLOR_DARK_TEXT     = "FF000000"  # 검정 텍스트
COLOR_WHITE_TEXT    = "FFFFFFFF"  # 흰색 텍스트


def _make_header_fill(color: str) -> PatternFill:
    """지정 색상의 셀 배경(채우기)을 만듭니다."""
    return PatternFill(fill_type="solid", fgColor=color)


def _make_border() -> Border:
    """셀 테두리 스타일을 만듭니다."""
    thin = Side(style="thin", color="FF000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_font(bold: bool = True, color: str = COLOR_WHITE_TEXT) -> Font:
    """헤더 셀에 사용할 글꼴을 만듭니다."""
    return Font(bold=bold, color=color, name="맑은 고딕", size=10)


def _normal_font() -> Font:
    """일반 데이터 셀에 사용할 글꼴을 만듭니다."""
    return Font(name="맑은 고딕", size=9)


def _write_header_row(
    ws,
    columns: List[str],
    bg_color: str = COLOR_HEADER_BLUE,
) -> None:
    """
    워크시트(ws)의 첫 행에 헤더(열 제목)를 씁니다.

    매개변수:
        ws      : 대상 워크시트
        columns : 헤더 문자열 목록
        bg_color: 배경 색상 코드
    """
    fill   = _make_header_fill(bg_color)
    font   = _header_font()
    border = _make_border()
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = fill
        cell.font   = font
        cell.border = border
        cell.alignment = align

    # 첫 행 높이 설정
    ws.row_dimensions[1].height = 20


def _write_data_row(
    ws,
    row_idx: int,
    values: List,
    fill_color: Optional[str] = None,
) -> None:
    """
    데이터 한 행을 워크시트에 씁니다.

    매개변수:
        ws         : 대상 워크시트
        row_idx    : 행 번호 (1-based)
        values     : 각 셀에 넣을 값 목록
        fill_color : 배경 색상 (없으면 흰색)
    """
    fill   = _make_header_fill(fill_color or COLOR_WHITE)
    font   = _normal_font()
    border = _make_border()
    align  = Alignment(vertical="top", wrap_text=True)

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill      = fill
        cell.font      = font
        cell.border    = border
        cell.alignment = align


def _auto_width(ws, max_width: int = 60) -> None:
    """
    각 열의 너비를 내용에 맞게 자동 조절합니다.
    너무 넓어지지 않도록 max_width로 제한합니다.
    """
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # 한글은 영문보다 약 2배 너비를 차지하므로 조정합니다
                text = str(cell.value)
                estimated = sum(2 if ord(c) > 127 else 1 for c in text)
                max_len = max(max_len, estimated)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _fmt_dt(dt: Optional[datetime]) -> str:
    """datetime 객체를 '2026-03-23 14:01:01' 형식 문자열로 변환합니다."""
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ============================================================
# 메인 리포터 클래스
# ============================================================

class ExcelReporter:
    """
    분석 결과를 받아서 Excel 파일(.xlsx)로 저장하는 클래스입니다.

    사용 예시:
        reporter = ExcelReporter()
        reporter.save(
            output_path  = "결과.xlsx",
            total_entries= 50000,
            spike_records= [...],
            error_records= [...],
            ...
        )
    """

    def save(
        self,
        output_path:       str | Path,
        total_entries:     int,
        parse_stats:       dict,
        spike_records:     List[SpikeRecord],
        error_records:     List[ErrorRecord],
        response_records:  List[ResponseTimeRecord],
        access_records:    List[AbnormalAccessRecord],
        wallet_records:    List[WalletAnomalyRecord],
        app_flow_records:  List[ApplicationFlowRecord],
    ) -> Path:
        """
        모든 분석 결과를 Excel 파일로 저장합니다.

        매개변수:
            output_path      : 저장할 파일 경로
            total_entries    : 분석한 총 로그 건수
            parse_stats      : 파싱 통계 딕셔너리
            spike_records    : 요청 폭증 기록
            error_records    : 에러 패턴 기록
            response_records : 응답시간 지연 기록
            access_records   : 비정상 접근 기록
            wallet_records   : 전자지갑 이상 기록
            app_flow_records : 원서접수 이상 기록

        반환:
            저장된 파일의 Path 객체
        """
        wb = Workbook()

        # 기본으로 생성되는 빈 시트를 제거합니다
        wb.remove(wb.active)

        # ---- 1. 분석요약 시트 ----
        self._write_summary_sheet(
            wb,
            total_entries,
            parse_stats,
            spike_records,
            error_records,
            response_records,
            access_records,
            wallet_records,
            app_flow_records,
        )

        # ---- 2. 요청폭증 시트 ----
        self._write_spike_sheet(wb, spike_records)

        # ---- 3. 에러패턴 시트 ----
        self._write_error_sheet(wb, error_records)

        # ---- 4. 응답시간지연 시트 ----
        self._write_response_time_sheet(wb, response_records)

        # ---- 5. 비정상접근 시트 ----
        self._write_access_sheet(wb, access_records)

        # ---- 6. 전자지갑이상 시트 ----
        self._write_wallet_sheet(wb, wallet_records)

        # ---- 7. 원서접수이상 시트 ----
        self._write_app_flow_sheet(wb, app_flow_records)

        # 파일 저장
        out = Path(output_path)
        wb.save(str(out))
        return out

    # ----------------------------------------------------------
    # 시트별 작성 메서드
    # ----------------------------------------------------------

    def _write_summary_sheet(
        self,
        wb: Workbook,
        total_entries: int,
        parse_stats: dict,
        spike_records,
        error_records,
        response_records,
        access_records,
        wallet_records,
        app_flow_records,
    ) -> None:
        """분석요약 시트를 작성합니다."""
        ws = wb.create_sheet("분석요약")

        # ---- 행 1: 큰 제목 (A1:C1 병합) ----
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value     = "JEUS 서버 로그 분석 결과 요약"
        title_cell.font      = Font(bold=True, size=14, name="맑은 고딕",
                                    color=COLOR_WHITE_TEXT)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill      = _make_header_fill(COLOR_HEADER_BLUE)
        ws.row_dimensions[1].height = 30

        # ---- 행 2: 분석 일시 ----
        ws["A2"] = "분석 일시"
        ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A2"].font = _normal_font()
        ws["B2"].font = _normal_font()

        # ---- 행 3: 빈 구분 행 ----
        # (아무것도 쓰지 않습니다)

        # ---- 행 4: 헤더 / 행 5~: 데이터 ----
        HEADER_ROW = 4   # 헤더를 4행에 씁니다 (1~3행은 제목/날짜/빈행)
        DATA_START  = 5  # 데이터를 5행부터 씁니다

        # 헤더 행 작성 (행 번호를 직접 지정)
        header_cols = ["항목", "수치", "비고"]
        fill   = _make_header_fill(COLOR_HEADER_BLUE)
        font   = _header_font()
        border = _make_border()
        align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, col_name in enumerate(header_cols, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
            cell.fill      = fill
            cell.font      = font
            cell.border    = border
            cell.alignment = align
        ws.row_dimensions[HEADER_ROW].height = 20

        # 항목별 요약 데이터 (헤더 제외)
        summary_data = [
            ("전체 로그 건수",      total_entries,                   "파싱된 유효 로그"),
            ("JEUS 시스템 로그",   parse_stats.get("jeus",    0),   ""),
            ("애플리케이션 로그",   parse_stats.get("egov",    0),   ""),
            ("인식 불가 줄",        parse_stats.get("unknown", 0),   "두 형식 모두 해당 없음"),
            ("",                   "",                               ""),
            ("요청 폭증 건수",      len(spike_records),              "임계값(3배) 초과 구간"),
            ("에러 패턴 클래스 수", len(error_records),              "ERROR/FATAL 발생 클래스"),
            ("응답시간 지연 건수",  len(response_records),           "3초 초과 처리"),
            ("비정상 접근 건수",    len(access_records),             "5분 내 30회 초과"),
            ("전자지갑 이상 건수",  len(wallet_records),             "선행 생성 없는 접근"),
            ("원서접수 이상 건수",  len(app_flow_records),           "오류 키워드/레벨 탐지"),
            ("",                   "",                               ""),
            ("총 이상현상 건수",
             sum([len(spike_records), len(error_records), len(response_records),
                  len(access_records), len(wallet_records), len(app_flow_records)]),
             "6개 항목 합계"),
        ]

        # 데이터 행 작성
        total_data_rows = len(summary_data)
        for row_offset, row_data in enumerate(summary_data):
            row_idx = DATA_START + row_offset
            if row_data[0] == "":
                continue  # 빈 구분 행은 건너뜁니다
            # 마지막 합계 행은 강조 색상 사용
            is_last = (row_offset == total_data_rows - 1)
            fill_color = COLOR_ORANGE_FILL if is_last else None
            _write_data_row(ws, row_idx, list(row_data), fill_color)

        # 열 너비 자동 조절
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 30

    def _write_spike_sheet(
        self, wb: Workbook, records: List[SpikeRecord]
    ) -> None:
        """요청 폭증 시트를 작성합니다."""
        ws = wb.create_sheet("요청폭증")

        headers = ["시간 구간", "시간 단위", "요청 수", "이전 대비 증가율", "비고"]
        _write_header_row(ws, headers, COLOR_HEADER_BLUE)

        if not records:
            ws.cell(row=2, column=1, value="탐지된 요청 폭증 없음")
            _auto_width(ws)
            return

        for i, rec in enumerate(records, start=2):
            # 증가율이 클수록 더 강한 색상으로 표시
            fill = COLOR_RED_FILL if rec.increase_rate >= 5 else COLOR_ORANGE_FILL
            _write_data_row(ws, i, [
                rec.window_label,
                rec.window_type,
                rec.request_count,
                f"{rec.increase_rate:.1f}배",
                "심각" if rec.increase_rate >= 5 else "경고",
            ], fill)

        _auto_width(ws)

    def _write_error_sheet(
        self, wb: Workbook, records: List[ErrorRecord]
    ) -> None:
        """에러 패턴 시트를 작성합니다."""
        ws = wb.create_sheet("에러패턴")

        headers = ["클래스명", "에러 건수", "최초 발생", "최근 발생",
                   "에러 메시지 샘플1", "에러 메시지 샘플2", "에러 메시지 샘플3"]
        _write_header_row(ws, headers, "FFCC0000")  # 빨간 헤더

        if not records:
            ws.cell(row=2, column=1, value="탐지된 에러 패턴 없음")
            _auto_width(ws)
            return

        for i, rec in enumerate(records, start=2):
            fill = COLOR_RED_FILL if rec.error_count >= 10 else COLOR_ORANGE_FILL
            samples = rec.sample_messages + ["", "", ""]  # 3개 미만이면 빈 문자열 채움
            _write_data_row(ws, i, [
                rec.source_class,
                rec.error_count,
                _fmt_dt(rec.first_occurred),
                _fmt_dt(rec.last_occurred),
                samples[0],
                samples[1],
                samples[2],
            ], fill)

        _auto_width(ws)

    def _write_response_time_sheet(
        self, wb: Workbook, records: List[ResponseTimeRecord]
    ) -> None:
        """응답시간 지연 시트를 작성합니다."""
        ws = wb.create_sheet("응답시간지연")

        headers = ["메서드명", "처리시간(ms)", "처리시간(초)", "시작 시각",
                   "종료 시각", "지연 여부", "시작 로그 (요약)", "종료 로그 (요약)"]
        _write_header_row(ws, headers, "FF7030A0")  # 보라 헤더

        if not records:
            ws.cell(row=2, column=1, value="탐지된 응답시간 지연 없음")
            _auto_width(ws)
            return

        for i, rec in enumerate(records, start=2):
            fill = COLOR_RED_FILL if rec.is_slow else COLOR_ORANGE_FILL
            _write_data_row(ws, i, [
                rec.method_name,
                round(rec.duration_ms, 1),
                round(rec.duration_ms / 1000, 2),
                _fmt_dt(rec.start_time),
                _fmt_dt(rec.end_time),
                "지연" if rec.is_slow else "정상",
                rec.start_log[:100],
                rec.end_log[:100],
            ], fill)

        _auto_width(ws)

    def _write_access_sheet(
        self, wb: Workbook, records: List[AbnormalAccessRecord]
    ) -> None:
        """비정상 접근 시트를 작성합니다."""
        ws = wb.create_sheet("비정상접근")

        headers = ["사용자ID / IP", "요청 횟수", "시작 시각", "종료 시각",
                   "요청 샘플1", "요청 샘플2", "요청 샘플3"]
        _write_header_row(ws, headers, "FFED7D31")  # 주황 헤더

        if not records:
            ws.cell(row=2, column=1, value="탐지된 비정상 접근 없음")
            _auto_width(ws)
            return

        for i, rec in enumerate(records, start=2):
            fill = COLOR_RED_FILL if rec.request_count >= 50 else COLOR_ORANGE_FILL
            samples = rec.sample_messages + ["", "", ""]
            _write_data_row(ws, i, [
                rec.identifier,
                rec.request_count,
                _fmt_dt(rec.time_range_start),
                _fmt_dt(rec.time_range_end),
                samples[0][:100],
                samples[1][:100] if len(samples) > 1 else "",
                samples[2][:100] if len(samples) > 2 else "",
            ], fill)

        _auto_width(ws)

    def _write_wallet_sheet(
        self, wb: Workbook, records: List[WalletAnomalyRecord]
    ) -> None:
        """전자지갑 이상 시트를 작성합니다."""
        ws = wb.create_sheet("전자지갑이상")

        headers = ["사용자 ID", "접근 시각", "접근 클래스/URL", "이상 내용", "원본 로그"]
        _write_header_row(ws, headers, "FF00B0F0")  # 하늘색 헤더

        if not records:
            ws.cell(row=2, column=1, value="탐지된 전자지갑 이상 없음")
            _auto_width(ws)
            return

        for i, rec in enumerate(records, start=2):
            _write_data_row(ws, i, [
                rec.user_id,
                _fmt_dt(rec.access_time),
                rec.access_class,
                rec.anomaly_desc,
                rec.raw_log[:200],
            ], COLOR_RED_FILL)

        _auto_width(ws)

    def _write_app_flow_sheet(
        self, wb: Workbook, records: List[ApplicationFlowRecord]
    ) -> None:
        """원서접수 이상 시트를 작성합니다."""
        ws = wb.create_sheet("원서접수이상")

        headers = ["발생 시각", "관련 클래스", "이상 내용", "원본 로그"]
        _write_header_row(ws, headers, "FF70AD47")  # 초록 헤더

        if not records:
            ws.cell(row=2, column=1, value="탐지된 원서접수 이상 없음")
            _auto_width(ws)
            return

        for i, rec in enumerate(records, start=2):
            _write_data_row(ws, i, [
                _fmt_dt(rec.occurred_at),
                rec.related_class,
                rec.anomaly_desc,
                rec.raw_log[:200],
            ], COLOR_ORANGE_FILL)

        _auto_width(ws)
